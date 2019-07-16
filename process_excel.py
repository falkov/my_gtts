import os
# import sys
# import time
# import requests
# import lxml.html as lxhtml
import logging
import openpyxl
import re


logging.basicConfig(format="%(asctime)s  %(filename)s:%(lineno)d  %(message)s",
                    datefmt="%Y-%m-%d  %H:%M.%S", level=logging.DEBUG,
                    filename=f'{os.path.dirname(__file__)}/_mylog.log')


class ExcelReader:
    def __init__(self, excelfile_name, sheet_name, row_from, row_to):
        self.sentences = list()
        self.sheet_name = sheet_name
        self.row_from = row_from
        self.row_to = row_to
        self.excelfile_path = f'{os.path.dirname(__file__)}/{excelfile_name}'

        def check_input_params(self):
            if not os.path.exists(self.excelfile_path):
                logging.info(f'No file: "{self.excelfile_path}"')
                exit(1)

            try:
                book = openpyxl.load_workbook(filename=self.excelfile_path)
            except Exception as ex:
                logging.info(f'Error in "book = openpyxl.load_workbook(filename=self.excelfile_path)" - {ex}')
                exit(1)

            sheet_names = book.sheetnames

            if self.sheet_name not in sheet_names:
                logging.info(f'No sheet "{sheet_name}" in sheets:"{sheet_names}"')
                exit(1)

        check_input_params(self)


    def read_data_from_xls(self):
        sentences_for_return = list()

        book = openpyxl.load_workbook(filename=self.excelfile_path)
        sheet = book[self.sheet_name]

        def mysplit(string_for_split):
            split_delimiters = r'falkov_tochka', 'falkov_voskl', 'falkov_vopr', '...'
            split_pattern = '|'.join(map(re.escape, split_delimiters))
            delete_pattern = r'\([^()]*\)'  # удалить скобки и все что в них

            string_for_split = string_for_split.replace('-', '')
            string_for_split = re.sub(delete_pattern, '', string_for_split)
            string_for_split = string_for_split.replace('!', '! falkov_voskl')
            string_for_split = string_for_split.replace('?', '? falkov_vopr')
            string_for_split = string_for_split.replace('.', '. falkov_tochka')

            split_sentences = re.split(split_pattern, string_for_split)
            clean_sentences = list()

            for split_sentence in split_sentences:
                if len(split_sentence) > 10:
                    clean_sentences.append(split_sentence.strip())

            return clean_sentences

        for row in range(self.row_from, self.row_to):
            excel_value = str(sheet["A" + str(row)].value)
            sentences = mysplit(excel_value)

            for sentence in sentences:
                sentences_for_return.append(sentence)

        return sentences_for_return


def main():
    excelreader = ExcelReader('English.xlsx', 'Sheet3', 1, 305)
    sentences = excelreader.read_data_from_xls()

    with open('english_text.txt', 'w') as file:
        for sentence in sentences:
            file.write(f'{sentence}\n')
            # file.write('[[slnc 3000]]\n')

    # os.system(f'say -v Samantha "What have you done with my pen? [[slnc 1000]]{sentences[1]}" -o cli.mp4')
    # os.system(f'say -v Alex "What have you done with my pen? [[slnc 1000]]{sentences[1]}" -o cli.mp4')

if __name__ == "__main__":
    main()
